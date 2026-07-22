"""평가일·agency 선택·Excel 공개/관리자 분리·재무 요약 단위 테스트."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pandas as pd

from common.settings import get_instruments_config
from export.agency_select import (
    is_usable_financial,
    latest_evaluation_date,
    select_one_per_company,
    select_result_for_company,
)
from export.excel import (
    ADMIN_COLUMNS,
    EXCEL_PUBLIC_COLUMNS,
    build_admin_rows,
    build_excel_public_rows,
    write_admin_excel_bytes,
    write_results_excel_tmp,
)
from export.fin_excel_utils import (
    build_summary_periods,
    build_summary_row_specs,
    cascade_summary_rows,
    convert_to_eok,
    excel_number_format,
    format_raw_unit_caption,
    lookup_raw_value,
    normalize_fin_table_headers,
    normalize_period_label,
    resolve_funding_row,
    resolve_ratio_row,
    shared_unit_caption,
)
from export.rating_excel_utils import build_cascaded_rating_rows
from extract.evaluation_date import (
    evaluation_date_from_filename,
    evaluation_date_from_text,
    extract_evaluation_date,
)
from extract.fin_tables import (
    filter_financial_data_rows,
    repair_financial_row_label,
)
from extract.rating_from_grid import is_rating_tokens_only_label


def _usable_table(**kwargs) -> dict:
    """기간 2열 + 라벨 행 3개 이상."""
    base = {
        "headers": ["구분", "2024.12", "2025.12"],
        "rows": [
            ["총자산", "100", "110"],
            ["당기순이익", "1", "2"],
            ["자기자본", "10", "11"],
        ],
        "unit_caption": "단위: 십억원, %",
    }
    base.update(kwargs)
    return base


def test_evaluation_date_from_text_label() -> None:
    text = "등급확정일 : 2026.05.21\n기타"
    assert evaluation_date_from_text(text) == "2026.05.21"


def test_evaluation_date_from_text_slash() -> None:
    text = "발행일 2026/05/21"
    assert evaluation_date_from_text(text) == "2026.05.21"


def test_evaluation_date_from_filename() -> None:
    assert (
        evaluation_date_from_filename("company_rs20260521-nice.pdf")
        == "2026.05.21"
    )


def test_evaluation_date_filename_fallback_only() -> None:
    assert extract_evaluation_date(None, "x_rs20260102-a.pdf") == "2026.01.02"
    assert extract_evaluation_date("본문만", "no_date.pdf") is None


def test_normalize_period_label_cases() -> None:
    assert normalize_period_label("2025") == "2025.12"
    assert normalize_period_label("2025(12)") == "2025.12"
    assert normalize_period_label("2025.12") == "2025.12"
    assert normalize_period_label("2026.03") == "2026.03"
    assert normalize_period_label("2026(03)") == "2026.03"


def test_agency_select_prefers_fin_over_higher_agency() -> None:
    nice = {
        "company_name": "테스트은행",
        "agency": "NICE신용평가㈜",
        "products": [{"status": "success", "rating": "AA"}],
        "financial_tables": [],
    }
    kis = {
        "company_name": "테스트은행",
        "agency": "한국신용평가㈜",
        "products": [{"status": "success", "rating": "AA-"}],
        "financial_tables": [_usable_table()],
    }
    assert select_result_for_company([kis, nice]) is kis


def test_agency_select_rejects_sparse_header_fin() -> None:
    """P1: 빈 헤더가 많은 KIS 표는 usable 아님 → KR 선택."""
    kis = {
        "company_name": "한국전력공사",
        "agency": "한국신용평가㈜",
        "products": [{"status": "success", "rating": "AAA"}],
        "financial_tables": [
            {
                "headers": [
                    "",
                    "",
                    "",
                    "2022.12",
                    "",
                    "2023.12",
                    "",
                    "2024.12",
                ],
                "rows": [
                    ["매출액(억원)", "", "", "1", "", "2", "", "3"],
                    ["영업이익(억원)", "", "", "1", "", "2", "", "3"],
                    ["당기순이익(억원)", "", "", "1", "", "2", "", "3"],
                ],
            }
        ],
    }
    kr = {
        "company_name": "한국전력공사",
        "agency": "한국기업평가㈜",
        "products": [{"status": "success", "rating": "AAA"}],
        "financial_tables": [_usable_table()],
    }
    assert not is_usable_financial(kis)
    assert select_result_for_company([kis, kr]) is kr


def test_agency_select_nice_when_both_usable() -> None:
    nice = {
        "company_name": "테스트은행",
        "agency": "NICE신용평가㈜",
        "products": [{"status": "fail", "rating": None}],
        "financial_tables": [_usable_table(rows=[
            ["총자산", "2", "3"],
            ["당기순이익", "1", "1"],
            ["자기자본", "1", "1"],
        ])],
    }
    kis = {
        "company_name": "테스트은행",
        "agency": "한국신용평가㈜",
        "products": [{"status": "success", "rating": "AA-"}],
        "financial_tables": [_usable_table()],
    }
    assert select_result_for_company([kis, nice]) is nice


def test_agency_select_kis_when_only_fin() -> None:
    kis = {
        "company_name": "테스트은행",
        "agency": "한국신용평가㈜",
        "products": [{"status": "fail", "rating": None}],
        "financial_tables": [_usable_table()],
    }
    kr = {
        "company_name": "테스트은행",
        "agency": "한국기업평가㈜",
        "products": [{"status": "success", "rating": "A"}],
        "financial_tables": [],
    }
    assert select_result_for_company([kr, kis]) is kis


def test_select_one_per_company_fallback_no_fin() -> None:
    results = [
        {
            "company_name": "A은행",
            "agency": "한국신용평가㈜",
            "products": [{"status": "success", "rating": "A"}],
            "financial_tables": [],
        },
        {
            "company_name": "A은행",
            "agency": "NICE신용평가㈜",
            "products": [{"status": "success", "rating": "AA"}],
            "financial_tables": [],
        },
        {
            "company_name": "B은행",
            "agency": "한국기업평가㈜",
            "products": [{"status": "success", "rating": "BBB"}],
            "financial_tables": [],
        },
    ]
    selected = select_one_per_company(results)
    assert len(selected) == 2
    assert selected[0]["agency"].startswith("NICE")


def test_summary_periods_year_end() -> None:
    assert build_summary_periods(["2024.12", "2025.12", "2026.03"]) == [
        "2024.12",
        "2025.12",
        "2026.03",
    ]


def test_summary_periods_reits_fallback() -> None:
    periods = [
        "2024.07",
        "2024.10",
        "2025.01",
        "2025.04",
        "2025.07",
        "2025.10",
    ]
    assert build_summary_periods(periods) == [
        "2025.04",
        "2025.07",
        "2025.10",
    ]


def test_unit_caption_shared_no_eok_convert() -> None:
    assert format_raw_unit_caption("단위: 십억원, %") == "(단위:십억,%)"
    assert shared_unit_caption({"unit_caption": "단위: 억원, %"}) == "(단위:억,%)"
    assert shared_unit_caption(
        {"unit_caption": None, "rows": [["총자산(억원)", "1"]]}
    ) == "(단위:억,%)"
    assert convert_to_eok(100.0, "십억") == 100.0


def test_excel_number_format() -> None:
    assert excel_number_format("1,234") == "#,##0"
    assert excel_number_format("12.5") == "#,##0.0"
    assert excel_number_format("1.0") == "#,##0.0"


def test_lookup_no_conversion() -> None:
    table = {
        "unit_caption": "단위: 십억원, %",
        "headers": ["구분", "2024.12", "2025.12", "2026.03"],
        "rows": [
            ["총자산", "100", "110", "120"],
            ["부채비율(%)", "50.1", "55.2", "60.3"],
        ],
    }
    value, raw = lookup_raw_value(
        table, metric_key="total_assets", period="2026.03"
    )
    assert value == 120.0
    assert raw == "120"
    debt, debt_raw = lookup_raw_value(
        table, metric_key="debt_ratio", period="2026.03"
    )
    assert debt == 60.3
    assert debt_raw == "60.3"


def test_resolve_ratio_debt_over_bis() -> None:
    table = {
        "headers": ["구분", "2024.12", "2025.12"],
        "rows": [
            ["총자산", "1", "2"],
            ["BIS기준 총자본비율", "14.1", "14.5"],
            ["부채비율(별도기준)", "30", "31"],
        ],
    }
    display, key = resolve_ratio_row(table)
    assert display == "부채비율(%)"
    assert key == "debt_ratio"


def test_resolve_ratio_bis_when_no_debt() -> None:
    table = {
        "headers": ["구분", "2024.12", "2025.12"],
        "rows": [
            ["총자산", "1", "2"],
            ["BIS기준 총자본비율", "14.1", "14.5"],
        ],
    }
    display, key = resolve_ratio_row(table)
    assert display == "BIS자본비율(%)"
    assert key == "bis_ratio"


def test_resolve_ratio_liquidity() -> None:
    table = {
        "headers": ["구분", "2024.12", "2025.12"],
        "rows": [
            ["총자산", "1", "2"],
            ["유동성비율", "150", "160"],
        ],
    }
    display, key = resolve_ratio_row(table)
    assert display == "유동성비율(%)"
    assert key == "liquidity_ratio"


def test_resolve_ratio_debt_default() -> None:
    table = {
        "headers": ["구분", "2024.12", "2025.12"],
        "rows": [
            ["총자산", "1", "2"],
            ["부채비율(%)", "80", "90"],
        ],
    }
    display, key = resolve_ratio_row(table)
    assert display == "부채비율(%)"
    assert key == "debt_ratio"
    specs = build_summary_row_specs(table)
    assert specs[-1] == ("부채비율(%)", "debt_ratio")


def test_resolve_funding_equity_fallback() -> None:
    table = {
        "headers": ["구분", "2024.12", "2025.12"],
        "rows": [
            ["총자산", "1", "2"],
            ["자기자본", "10", "11"],
            ["당기순이익", "1", "1"],
        ],
    }
    display, key = resolve_funding_row(table)
    assert display == "자기자본"
    assert key == "equity"
    specs = build_summary_row_specs(table)
    assert specs[2] == ("자기자본", "equity")


def test_resolve_ratio_leverage_fallback() -> None:
    table = {
        "headers": ["구분", "2024.12", "2025.12"],
        "rows": [
            ["총자산", "1", "2"],
            ["총자산/자기자본(배)", "8", "9"],
        ],
    }
    display, key = resolve_ratio_row(table)
    assert display == "총자산/자기자본(배)"
    assert key == "leverage"


def test_net_income_exact_no_substring() -> None:
    table = {
        "headers": ["구분", "2024.12", "2025.12"],
        "rows": [
            ["총자산", "1", "2"],
            ["지배주주지분순이익", "10", "11"],
        ],
    }
    value, _ = lookup_raw_value(
        table, metric_key="net_income", period="2025.12"
    )
    assert value is None
    specs = build_summary_row_specs(table)
    assert specs[1] == ("당기순이익", "net_income")


def test_summary_defaults_when_empty() -> None:
    specs = build_summary_row_specs(
        {"headers": ["구분", "2024.12"], "rows": [["기타", "1"]]}
    )
    assert specs == [
        ("총자산", "total_assets"),
        ("당기순이익", "net_income"),
        ("총차입금", "total_borrowings"),
        ("부채비율(%)", "debt_ratio"),
    ]


def test_repair_financial_row_label_trailing_number() -> None:
    row = repair_financial_row_label(
        ["총자산 3,331,525", "3331525", "3597145"]
    )
    assert row[0] == "총자산"
    assert row[1] == "3,331,525"

    concat_row = repair_financial_row_label(
        ["당기순이익 20,925", "20314212111237", "21112372348050"]
    )
    assert concat_row[0] == "당기순이익"
    assert concat_row[1] == "20,925"


def test_equity_alias_capital_total() -> None:
    from common.metric_catalog import get_metrics_config

    lookup = get_metrics_config().normalized_lookup
    assert lookup.get("자본총계") == "equity"


def test_cascade_summary_fills_empty_row_from_next_agency() -> None:
    nice = {
        "headers": ["구분", "2023.12", "2024.12", "2025.12"],
        "rows": [
            ["총자산", "1", "2", "3"],
            ["당기순이익", "1", "1", "1"],
            ["자기자본", "10", "11", "12"],
        ],
    }
    kr = {
        "headers": ["구분", "2023.12", "2024.12", "2025.12"],
        "rows": [
            ["총자산", "1", "2", "3"],
            ["당기순이익", "1", "1", "1"],
            ["자기자본", "10", "11", "12"],
            ["부채비율(%)", "80", "85", "90"],
        ],
    }
    periods, _unit, rows = cascade_summary_rows([nice, kr])
    assert periods == ["2023.12", "2024.12", "2025.12"]
    assert rows[3][0] == "부채비율(%)"
    assert rows[3][2][-1][0] == 90.0


def test_cascaded_rating_rows_union() -> None:
    config = get_instruments_config()
    results = [
        {
            "agency": "NICE신용평가㈜",
            "products": [
                {
                    "instrument_key": "issuer",
                    "raw_label": "Issuer",
                    "rating": "AA+",
                    "status": "success",
                },
                {
                    "instrument_key": "senior_unsecured",
                    "raw_label": "Bond",
                    "rating": "AA",
                    "status": "success",
                },
            ],
        },
        {
            "agency": "한국기업평가㈜",
            "products": [
                {
                    "instrument_key": "coco_t1",
                    "raw_label": "CoCo",
                    "rating": "A+",
                    "status": "success",
                },
            ],
        },
    ]
    rows = build_cascaded_rating_rows(results, config)
    labels = [label for label, _rating in rows]
    assert "발행자신용등급" in labels
    assert "무보증사채" in labels
    assert "조건부자본증권(신종)" in labels


def test_latest_evaluation_date() -> None:
    results = [
        {"evaluation_date": "2026.01.15"},
        {"evaluation_date": "2026.05.21"},
        {"evaluation_date": "2026.03.01"},
    ]
    assert latest_evaluation_date(results) == "2026.05.21"


def test_is_rating_tokens_only_label() -> None:
    assert is_rating_tokens_only_label("A+ A A-")
    assert not is_rating_tokens_only_label("무보증사채")


def test_filter_financial_narrative_rows() -> None:
    rows = [
        ["총자산", "1", "2"],
        ["적용재무제표", "연결", "연결"],
        ["평정 논거", "", ""],
        ["\uf06c 우수한 시장지위", "", ""],
        ["2. BIS기준 총자본비율은", "", ""],
    ]
    filtered = filter_financial_data_rows(rows)
    assert [r[0] for r in filtered] == ["총자산", "적용재무제표"]


def test_normalize_drops_empty_period_cols() -> None:
    table = {
        "headers": ["", "", "", "2022.12", "", "2023.12"],
        "rows": [["총자산(억원)", "", "", "100", "", "110"]],
    }
    headers, rows = normalize_fin_table_headers(table)
    assert headers == ["구분", "2022.12", "2023.12"]
    assert rows[0] == ["총자산(억원)", "100", "110"]


def test_public_vs_admin_columns() -> None:
    config = get_instruments_config()
    result = {
        "result_no": 1,
        "company_name": "테스트은행",
        "agency": "NICE신용평가㈜",
        "evaluation_date": "2026.05.21",
        "file_name": "a.pdf",
        "status": "partial",
        "products": [
            {
                "instrument_key": "issuer",
                "raw_label": "Issuer Rating",
                "rating": "AAA",
                "outlook": "안정적",
                "evaluation_type": "정기",
                "status": "success",
                "fail_reason": None,
            },
            {
                "instrument_key": "coco_t1",
                "raw_label": "신종자본증권",
                "rating": None,
                "outlook": None,
                "evaluation_type": "본",
                "status": "fail",
                "fail_reason": {"code": "rating_not_found"},
            },
        ],
        "fail_reason": None,
    }
    public = build_excel_public_rows(result, config)
    admin = build_admin_rows(result, config)
    assert list(public[0].keys()) == EXCEL_PUBLIC_COLUMNS
    assert len(public) == 1
    assert len(admin) == 2


def test_company_sheet_summary_layout(tmp_path: Path) -> None:
    config = get_instruments_config()
    result = {
        "result_no": 1,
        "company_name": "요약은행",
        "agency": "NICE신용평가㈜",
        "evaluation_date": "2026.01.01",
        "file_name": "s.pdf",
        "status": "success",
        "products": [
            {
                "instrument_key": "issuer",
                "raw_label": "Issuer",
                "rating": "AA",
                "outlook": "안정적",
                "evaluation_type": "정기",
                "status": "success",
                "fail_reason": None,
            }
        ],
        "financial_tables": [
            {
                "unit_caption": "단위: 십억원, %",
                "headers": ["구분", "2024.12", "2025.12", "2026.03"],
                "rows": [
                    ["총자산", "100", "110", "120"],
                    ["당기순이익", "1", "2", "3"],
                    ["총차입금", "10", "11", "12"],
                    ["BIS자기자본비율(%)", "14.1", "14.2", "14.3"],
                    ["부채비율(%)", "40", "41", "42"],
                ],
            }
        ],
    }
    out = tmp_path / "out.xlsx"
    tmp = write_results_excel_tmp([result], config, out)
    xlsx_path = tmp.with_suffix(".xlsx")
    tmp.rename(xlsx_path)
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    ws = wb["요약은행"]
    values = [[cell.value for cell in row] for row in ws.iter_rows(max_col=4)]
    section_titles = [row[0] for row in values if row and row[0]]
    flat = [v for row in values for v in row]
    overview_idx = section_titles.index("개요")
    rating_summary_idx = section_titles.index("신용등급(요약)")
    fin_summary_idx = section_titles.index("재무지표(요약)")
    rating_raw_idx = section_titles.index("신용등급(원본)")
    fin_raw_idx = section_titles.index("재무지표(원본)")
    assert overview_idx < rating_summary_idx < fin_summary_idx
    assert fin_summary_idx < rating_raw_idx < fin_raw_idx
    assert "NICE신용평가㈜" in section_titles
    assert "신용등급(나이스" not in flat
    assert "재무지표(나이스" not in flat

    assert "재무비율" in flat
    assert "(단위:십억,%)" in flat
    assert "부채비율(%)" in flat
    assert "신평사" not in flat
    assert "원본파일명" not in flat

    summary_title = next(r for r in values if r and r[0] == "재무지표(요약)")
    assert "(단위:십억,%)" in summary_title
    summary_labels = []
    for i, row in enumerate(values):
        if row and row[0] == "재무비율":
            for later in values[i + 1 : i + 5]:
                if later and later[0]:
                    summary_labels.append(later[0])
            break
    assert summary_labels == [
        "총자산",
        "당기순이익",
        "총차입금",
        "부채비율(%)",
    ]


def test_admin_excel_single_sheet_columns() -> None:
    config = get_instruments_config()
    result = {
        "result_no": 1,
        "company_name": "관리자은행",
        "agency": "NICE신용평가㈜",
        "evaluation_date": "2026.05.21",
        "file_name": "a.pdf",
        "status": "success",
        "products": [
            {
                "instrument_key": "issuer",
                "raw_label": "Issuer",
                "rating": "AAA",
                "outlook": "안정적",
                "evaluation_type": "정기",
                "status": "success",
                "fail_reason": None,
            }
        ],
    }
    payload = write_admin_excel_bytes([result], config)
    dataframe = pd.read_excel(BytesIO(payload))
    assert list(dataframe.columns) == ADMIN_COLUMNS
