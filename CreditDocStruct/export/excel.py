"""비개발자용 Excel(목록+기업시트) / 관리자용 행 빌더."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from common.settings import InstrumentsConfig
from export.agency_select import select_one_per_company
from export.fin_excel_utils import (
    build_summary_periods,
    build_summary_row_specs,
    excel_number_format,
    lookup_raw_value,
    normalize_fin_table_headers,
    shared_unit_caption,
)
from extract.fin_tables import parse_numeric_cell

EXCEL_PUBLIC_COLUMNS = [
    "No",
    "회사명",
    "신평사",
    "상품분류",
    "신용등급",
    "등급전망",
    "평가일",
    "원본파일명",
]

ADMIN_COLUMNS = [
    "No",
    "회사명",
    "신평사",
    "상품분류",
    "상품분류_Key",
    "원본라벨",
    "평가종류",
    "신용등급",
    "등급전망",
    "평가일",
    "원본파일명",
    "처리상태",
    "실패사유",
]

# 하위 호환: 기존 코드/테스트는 ADMIN 컬럼을 가리킴
EXCEL_COLUMNS = ADMIN_COLUMNS

_HEADER_FILL = PatternFill("solid", fgColor="002060")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SECTION_FONT = Font(bold=True, size=12)
_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_SHEET_FORBIDDEN = re.compile(r'[\\/*?:\[\]]')


def build_admin_rows(
    result: dict[str, Any],
    config: InstrumentsConfig,
) -> list[dict[str, Any]]:
    """관리자/JSON용: 모든 product (+ 빈 products면 fail 행)."""
    products = result.get("products") or []
    base = {
        "No": result.get("result_no", ""),
        "회사명": result.get("company_name", ""),
        "신평사": result.get("agency", ""),
        "평가일": result.get("evaluation_date") or "",
        "원본파일명": result.get("file_name", ""),
    }

    if not products:
        fail_reason = result.get("fail_reason") or {}
        return [
            {
                **base,
                "처리상태": result.get("status") or "fail",
                "상품분류_Key": "",
                "상품분류": "",
                "원본라벨": "",
                "평가종류": "",
                "신용등급": "",
                "등급전망": "",
                "실패사유": fail_reason.get("code") or "",
            }
        ]

    rows: list[dict[str, Any]] = []
    for product in products:
        instrument_key = product.get("instrument_key")
        display_name = ""
        if instrument_key:
            definition = config.instruments.get(instrument_key)
            display_name = definition.display_name if definition else ""

        fail_reason = product.get("fail_reason") or {}
        rows.append(
            {
                **base,
                "처리상태": product.get("status") or result.get("status") or "",
                "상품분류_Key": instrument_key or "",
                "상품분류": display_name,
                "원본라벨": product.get("raw_label") or "",
                "평가종류": product.get("evaluation_type") or "",
                "신용등급": product.get("rating") or "",
                "등급전망": product.get("outlook") or "",
                "실패사유": fail_reason.get("code") or "",
            }
        )
    return rows


def build_excel_rows(
    result: dict[str, Any],
    config: InstrumentsConfig,
) -> list[dict[str, Any]]:
    """하위 호환 alias → admin rows."""
    return build_admin_rows(result, config)


def build_excel_public_rows(
    result: dict[str, Any],
    config: InstrumentsConfig,
) -> list[dict[str, Any]]:
    """비개발자 목록: success product만."""
    rows: list[dict[str, Any]] = []
    for admin_row in build_admin_rows(result, config):
        if admin_row.get("처리상태") != "success":
            continue
        if not admin_row.get("신용등급"):
            continue
        rows.append(
            {
                "No": admin_row.get("No", ""),
                "회사명": admin_row.get("회사명", ""),
                "신평사": admin_row.get("신평사", ""),
                "상품분류": admin_row.get("상품분류", ""),
                "신용등급": admin_row.get("신용등급", ""),
                "등급전망": admin_row.get("등급전망", ""),
                "평가일": admin_row.get("평가일", ""),
                "원본파일명": admin_row.get("원본파일명", ""),
            }
        )
    return rows


def build_excel_row(
    result: dict[str, Any],
    config: InstrumentsConfig,
) -> dict[str, Any]:
    rows = build_admin_rows(result, config)
    return rows[0]


def build_financial_excel_rows(
    result: dict[str, Any],
) -> list[dict[str, Any]]:
    """JSON/내부용 long-format (Excel 시트에는 쓰지 않음)."""
    from common.settings import get_metrics_config

    metrics = get_metrics_config()
    base = {
        "No": result.get("result_no", ""),
        "회사명": result.get("company_name", ""),
        "신평사": result.get("agency", ""),
        "원본파일명": result.get("file_name", ""),
    }
    facts = result.get("financial_facts") or []
    rows: list[dict[str, Any]] = []
    for fact in facts:
        metric_key = fact.get("metric_key")
        display = ""
        if metric_key and metric_key in metrics.metrics:
            display = metrics.metrics[metric_key].display_name
        rows.append(
            {
                **base,
                "지표_Key": metric_key or "",
                "지표명": display,
                "원본라벨": fact.get("raw_label") or "",
                "매칭상태": fact.get("classification_status") or "",
                "기간": fact.get("period") or "",
                "값": fact.get("value"),
                "원본값": fact.get("value_raw") or "",
                "단위": fact.get("unit") or "",
                "기준": fact.get("basis") or "",
            }
        )
    return rows


def _autosize(worksheet, *, max_width: int | None = 120) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        width = max(max_length + 8, 5)
        if max_width is not None:
            width = min(width, max_width)
        worksheet.column_dimensions[column_letter].width = width


def _safe_sheet_title(name: str, used: set[str]) -> str:
    cleaned = _SHEET_FORBIDDEN.sub("_", (name or "시트").strip()) or "시트"
    cleaned = cleaned[:31]
    candidate = cleaned
    suffix = 1
    while candidate in used or candidate.lower() == "신용등급".lower():
        tail = f"_{suffix}"
        candidate = (cleaned[: 31 - len(tail)] + tail)[:31]
        suffix += 1
    used.add(candidate)
    return candidate


def _style_header_row(ws, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _THIN


def _apply_borders(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = _THIN


def _set_company_col_widths(ws, last_col: int) -> None:
    ws.column_dimensions["A"].width = 25
    for col in range(2, max(last_col, 2) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15


def _success_products(result: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        p
        for p in (result.get("products") or [])
        if p.get("status") == "success" and p.get("rating")
    ]


def _write_company_sheet(
    wb: Workbook,
    result: dict[str, Any],
    config: InstrumentsConfig,
    used_titles: set[str],
) -> None:
    title = _safe_sheet_title(str(result.get("company_name") or "기업"), used_titles)
    ws = wb.create_sheet(title)
    row = 1

    # --- 개요 ---
    ws.cell(row=row, column=1, value="개요").font = _SECTION_FONT
    row += 1
    overview = [
        ("회사명", result.get("company_name") or ""),
        ("신평사", result.get("agency") or ""),
        ("평가일", result.get("evaluation_date") or ""),
        ("원본파일명", result.get("file_name") or ""),
    ]
    start = row
    for label, value in overview:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1
    _style_header_row(ws, start, 1, 1)
    for r in range(start, row):
        ws.cell(row=r, column=1).fill = _HEADER_FILL
        ws.cell(row=r, column=1).font = _HEADER_FONT
        ws.cell(row=r, column=1).border = _THIN
        ws.cell(row=r, column=2).border = _THIN
    row += 1

    # --- 신용등급 ---
    ws.cell(row=row, column=1, value="신용등급").font = _SECTION_FONT
    row += 1
    rating_header_row = row
    ws.cell(row=row, column=1, value="상품분류")
    ws.cell(row=row, column=2, value="신용등급")
    _style_header_row(ws, row, 1, 2)
    row += 1
    for product in _success_products(result):
        key = product.get("instrument_key")
        display = ""
        if key and key in config.instruments:
            display = config.instruments[key].display_name
        elif product.get("raw_label"):
            display = str(product.get("raw_label"))
        ws.cell(row=row, column=1, value=display)
        ws.cell(row=row, column=2, value=product.get("rating") or "")
        ws.cell(row=row, column=1).border = _THIN
        ws.cell(row=row, column=2).border = _THIN
        row += 1
    if row == rating_header_row + 1:
        ws.cell(row=row, column=1, value="")
        ws.cell(row=row, column=2, value="")
        _apply_borders(ws, row, row, 1, 2)
        row += 1
    row += 1

    # --- 재무지표 ---
    fin_title_row = row
    ws.cell(row=row, column=1, value="재무지표").font = _SECTION_FONT
    tables = result.get("financial_tables") or []
    table = next((t for t in tables if t.get("rows")), None)
    last_col = 2
    unit_caption = shared_unit_caption(table)
    if table:
        headers, data_rows = normalize_fin_table_headers(table)
        last_col = max(len(headers), 2)
        ws.cell(row=fin_title_row, column=last_col, value=unit_caption)
        ws.cell(row=fin_title_row, column=last_col).alignment = Alignment(
            horizontal="right"
        )
        row += 1
        for col, header in enumerate(headers, start=1):
            ws.cell(row=row, column=col, value=header)
        _style_header_row(ws, row, 1, last_col)
        row += 1
        for data in data_rows:
            for col in range(1, last_col + 1):
                raw = data[col - 1] if col - 1 < len(data) else None
                cell = ws.cell(row=row, column=col)
                if col == 1:
                    cell.value = raw
                else:
                    num, raw_text = parse_numeric_cell(
                        raw if raw is not None else None
                    )
                    if num is not None:
                        cell.value = num
                        cell.number_format = excel_number_format(
                            raw_text if raw_text is not None else raw
                        )
                    else:
                        cell.value = raw
                cell.border = _THIN
            row += 1
    else:
        row += 1

    row += 1

    # --- 재무지표(요약) ---
    summary_title_row = row
    ws.cell(row=row, column=1, value="재무지표(요약)").font = _SECTION_FONT
    summary_periods = build_summary_periods(
        list((table.get("headers") or [])[1:]) if table else []
    )
    summary_specs = build_summary_row_specs(table)
    end_col = max(1 + len(summary_periods), 4)
    ws.cell(row=summary_title_row, column=end_col, value=unit_caption)
    ws.cell(row=summary_title_row, column=end_col).alignment = Alignment(
        horizontal="right"
    )
    row += 1

    ws.cell(row=row, column=1, value="재무비율")
    for col, period in enumerate(summary_periods, start=2):
        ws.cell(row=row, column=col, value=period)
    _style_header_row(ws, row, 1, end_col)
    row += 1

    if table and summary_periods:
        for display_label, metric_key in summary_specs:
            ws.cell(row=row, column=1, value=display_label).border = _THIN
            for col, period in enumerate(summary_periods, start=2):
                value, raw_text = lookup_raw_value(
                    table,
                    metric_key=metric_key,
                    period=period,
                )
                cell = ws.cell(row=row, column=col)
                if value is not None:
                    cell.value = value
                    cell.number_format = excel_number_format(raw_text)
                cell.border = _THIN
            for col in range(2 + len(summary_periods), end_col + 1):
                ws.cell(row=row, column=col).border = _THIN
            row += 1
    else:
        for display_label, _key in summary_specs:
            ws.cell(row=row, column=1, value=display_label).border = _THIN
            for col in range(2, end_col + 1):
                ws.cell(row=row, column=col).border = _THIN
            row += 1

    _set_company_col_widths(ws, max(last_col, end_col))


def write_results_excel_tmp(
    results: list[dict[str, Any]],
    config: InstrumentsConfig,
    final_path: str | Path,
) -> Path:
    """비개발자 Excel: 신용등급 목록 + 기업별 시트."""
    final_path = Path(final_path)
    tmp_path = final_path.with_name(final_path.name + ".tmp")
    tmp_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    list_ws = wb.active
    list_ws.title = "신용등급"

    for col, name in enumerate(EXCEL_PUBLIC_COLUMNS, start=1):
        cell = list_ws.cell(row=1, column=col, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _THIN

    out_row = 2
    for result in results:
        for item in build_excel_public_rows(result, config):
            for col, name in enumerate(EXCEL_PUBLIC_COLUMNS, start=1):
                cell = list_ws.cell(row=out_row, column=col, value=item.get(name, ""))
                cell.border = _THIN
            out_row += 1

    list_ws.freeze_panes = "A2"
    if out_row > 2:
        list_ws.auto_filter.ref = list_ws.dimensions
    _autosize(list_ws, max_width=120)

    used_titles: set[str] = {"신용등급"}
    for selected in select_one_per_company(results):
        _write_company_sheet(wb, selected, config, used_titles)

    wb.save(tmp_path)
    return tmp_path


def write_admin_excel_bytes(
    results: list[dict[str, Any]],
    config: InstrumentsConfig,
) -> bytes:
    """관리자 다운로드: 상세 목록 1시트만."""
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "신용등급"

    for col, name in enumerate(ADMIN_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _THIN

    out_row = 2
    for result in results:
        for item in build_admin_rows(result, config):
            for col, name in enumerate(ADMIN_COLUMNS, start=1):
                cell = ws.cell(row=out_row, column=col, value=item.get(name, ""))
                cell.border = _THIN
            out_row += 1

    ws.freeze_panes = "A2"
    if out_row > 2:
        ws.auto_filter.ref = ws.dimensions
    _autosize(ws, max_width=None)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
